"""
Müşteri Excel template'ini cell_map ile doldurur.

cell_map yapısı (customer.excel_config_json içinde):
{
  "vat_mode": "exclusive" | "inclusive",
  "header": {
      "B4": "event_name",
      "C5": "check_in",
      ...
  },
  "data_block": {
      "start_row": 11,
      "end_anchor_text": "ARA TOPLAM",
      "sheet": "Sheet1",
      "columns": {
          "B": "service_name",
          "C": "notes",
          "D": "nights",
          "E": "qty",
          "F": "sale_price_eur",
          "H": "sale_price"
      }
  }
}

Formül kolonları (G, I, J vb.) template'in ilk dolu satırından otomatik tespit edilir.
"""
from __future__ import annotations

import io
import os
import re

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string, get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from .builder import SECTION_LABELS, SECTIONS_ORDER


# Bölüm ara toplam etiketleri
SECTION_SUBTOTAL_LABELS: dict[str, str] = {
    "accommodation": "Konaklama Ara Toplam",
    "meeting":       "Toplantı / Salon Ara Toplam",
    "fb":            "Yeme & İçme Ara Toplam",
    "teknik":        "Teknik Ekipman Ara Toplam",
    "dekor":         "Dekor / Süsleme Ara Toplam",
    "transfer":      "Transfer / Ulaşım Ara Toplam",
    "tasarim":       "Tasarım & Baskı Ara Toplam",
    "other":         "Diğer Hizmetler Ara Toplam",
}


# ── Stil sabitleri ─────────────────────────────────────────────────────────────
def _HDR_FONT(): return Font(bold=True, color="FFFFFF", size=10)
def _HDR_FILL(): return PatternFill(fill_type="solid", fgColor="1E5F8C")
def _SUB_FONT(): return Font(bold=True, color="1A3A5C", size=10)
def _SUB_FILL(): return PatternFill(fill_type="solid", fgColor="D0E8F5")
def _TOT_FONT(): return Font(bold=True, color="FFFFFF", size=11)
def _TOT_FILL(): return PatternFill(fill_type="solid", fgColor="1A3A5C")
def _DAT_FONT(): return Font(color="000000", size=10)


def _sf_sale(budget, currency: str) -> float:
    """Hizmet bedeli tutarını offer_currency cinsinden hesaplar."""
    pct = float(budget.service_fee_pct or 0)
    if not pct:
        return 0.0
    offer_rate = budget.rate_to_try(currency) or 1.0
    base = 0.0
    for r in budget.rows:
        if r.get("is_service_fee") or r.get("is_accommodation_tax"):
            continue
        sale      = float(r.get("sale_price", 0) or 0)
        qty       = float(r.get("qty",    1) or 1)
        nights    = float(r.get("nights", 1) or 1)
        row_cur   = (r.get("currency") or "TRY").upper()
        row_rate  = budget.rate_to_try(row_cur) or 1.0
        conv      = row_rate / offer_rate
        base     += sale * qty * nights * conv
    return round(base * pct / 100, 2)


# ── Header alan çözücüleri ─────────────────────────────────────────────────────
def _header_resolvers(budget, request, customer, creator) -> dict:
    req = request

    def _date(d):
        if not d:
            return ""
        if isinstance(d, str):
            try:
                from datetime import date as _date_cls
                d = _date_cls.fromisoformat(d[:10])
            except Exception:
                return d
        return d.strftime("%d.%m.%Y")

    cities = ""
    if req:
        cities = (", ".join(req.cities)
                  if getattr(req, "cities", None)
                  else getattr(req, "city", ""))

    return {
        "event_name":     (getattr(req, "event_name", None)  or budget.venue_name or ""),
        "ref_no":         (getattr(req, "request_no", None)  or ""),
        "check_in":       _date(getattr(req, "check_in", None)),
        "check_out":      _date(getattr(req, "check_out", None)),
        "venue_name":     (budget.venue_name or ""),
        "customer_name":  (getattr(customer, "name", None)
                           or getattr(req, "client_name", None)
                           or ""),
        "creator_name":   (f"{getattr(creator, 'name', '')} "
                           f"{getattr(creator, 'surname', '')}".strip()
                           if creator else ""),
        "eur_rate":       budget.rate_to_try("EUR"),
        "usd_rate":       budget.rate_to_try("USD"),
        "attendee_count": (getattr(req, "attendee_count", None) or ""),
        "city":           cities,
        # Servis bedeli
        "sf_pct":         float(budget.service_fee_pct or 0),
        "sf_sale":        _sf_sale(budget, currency),
        "sf_vat":         round(_sf_sale(budget, currency) * 0.20, 2),
        "sf_total":       round(_sf_sale(budget, currency) * 1.20, 2),
    }


# ── Satır alan çözücüleri ──────────────────────────────────────────────────────
def _row_value(field: str, row: dict, budget, currency: str) -> float | str:
    qty    = float(row.get("qty",    1) or 1)
    nights = float(row.get("nights", 1) or 1)
    sale   = float(row.get("sale_price", 0) or 0)
    vat    = float(row.get("vat_rate",   0) or 0)

    row_cur = (row.get("currency") or "TRY").upper()
    if row_cur != currency:
        row_rate   = budget.rate_to_try(row_cur) or 1.0
        offer_rate = budget.rate_to_try(currency) or 1.0
        sale = sale * row_rate / offer_rate

    def _to_cur(target: str) -> float:
        try_val = float(row.get("sale_price", 0) or 0) * (budget.rate_to_try(row_cur) or 1.0)
        t_rate  = budget.rate_to_try(target) or 1.0
        return try_val / t_rate

    match field:
        case "service_name":   return row.get("service_name", "")
        case "notes":          return row.get("notes", "")
        case "unit":           return row.get("unit", "Adet")
        case "qty":            return qty
        case "nights":         return nights
        case "vat_rate":       return vat
        case "vat_pct":        return vat / 100
        case "sale_price":     return round(sale, 2)
        case "sale_price_inc": return round(sale * (1 + vat / 100), 2)
        case "total_excl":     return round(sale * qty * nights, 2)
        case "total_incl":     return round(sale * (1 + vat / 100) * qty * nights, 2)
        case "sale_price_eur": return round(_to_cur("EUR"), 2)
        case "sale_price_usd": return round(_to_cur("USD"), 2)
        case "total_eur":      return round(_to_cur("EUR") * qty * nights, 2)
        case "total_usd":      return round(_to_cur("USD") * qty * nights, 2)
        case _:                return ""


# ── Formül şablonu çıkarıcı ────────────────────────────────────────────────────
def _extract_formula_templates(ws, start_row: int, written_cols: set, max_col: int) -> dict:
    """
    Template'deki veri satırlarından formül şablonlarını çıkarır.
    Yazan kolonlar (col_defs) hariç, formül içeren kolonları tespit eder.
    Satır numarasını {row} ile değiştirir.
    """
    formula_cols: dict[str, str] = {}
    for r_idx in range(start_row, min(start_row + 30, ws.max_row + 1)):
        found = False
        for c_idx in range(1, max_col + 1):
            col_letter = get_column_letter(c_idx)
            if col_letter.upper() in written_cols:
                continue
            if col_letter in formula_cols:
                continue
            try:
                cell = ws.cell(row=r_idx, column=c_idx)
            except Exception:
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                # Satır numarasını {row} ile değiştir
                pattern = r'(?<=[A-Za-z\$])(' + str(r_idx) + r')(?=[^0-9]|$)'
                tpl = re.sub(pattern, '{row}', cell.value)
                formula_cols[col_letter] = tpl
                found = True
        if found:
            break
    return formula_cols


# ── Güvenli hücre yazma ────────────────────────────────────────────────────────
def _safe_set(ws, row: int, col_letter: str, value,
              font=None, fill=None) -> None:
    try:
        ci   = column_index_from_string(col_letter.upper())
        cell = ws.cell(row=row, column=ci)
        cell.value = value
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
    except (AttributeError, Exception):
        pass


# ── Ana fill fonksiyonu (in-place) ────────────────────────────────────────────
def _fill_ws(ws, cell_map: dict, budget, request, customer, creator) -> None:
    """Worksheet'i cell_map ile doldurur: header, bölüm başlıkları, ara toplamlar, genel toplam."""
    currency = (budget.offer_currency or "TRY").upper()

    # 1. Header hücrelerini doldur
    header_vals = _header_resolvers(budget, request, customer, creator)
    for cell_addr, field_name in (cell_map.get("header") or {}).items():
        val = header_vals.get(field_name)
        if val is not None:
            try:
                ws[cell_addr.upper()] = val
                ws[cell_addr.upper()].font = Font(color="000000")
            except (AttributeError, Exception):
                pass

    data_block = cell_map.get("data_block")
    if not data_block:
        return

    start_row  = int(data_block.get("start_row", 1))
    col_defs   = data_block.get("columns", {})           # {letter: field_name}
    end_anchor = data_block.get("end_anchor_text")
    label_col  = "B"                                      # bölüm etiketleri için varsayılan kolon

    # 2. Anchor satırını bul
    anchor_row = None
    if end_anchor:
        for r_idx in range(start_row, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                v = ws.cell(row=r_idx, column=c_idx).value
                if v and isinstance(v, str) and end_anchor.lower() in v.lower():
                    anchor_row = r_idx
                    break
            if anchor_row:
                break

    # 3. Formül şablonlarını temizlemeden önce çıkar
    written_cols = {c.upper() for c in col_defs}
    formula_cols = _extract_formula_templates(
        ws, start_row, written_cols, ws.max_column
    )
    print(f"[FILLER] start_row={start_row} anchor={anchor_row} written={sorted(written_cols)} formula_cols={dict(formula_cols)}", flush=True)

    # 4. Veri alanını temizle (start_row → anchor+10 arası)
    clear_end = (anchor_row + 10) if anchor_row else (start_row + 100)
    for r_idx in range(start_row, min(clear_end + 1, ws.max_row + 1)):
        for c_idx in range(1, ws.max_column + 1):
            try:
                ws.cell(row=r_idx, column=c_idx).value = None
            except AttributeError:
                pass

    # 5. Bütçe satırlarını bölümlere ayır
    rows_by_sec: dict[str, list] = {}
    service_fee = None
    for r in budget.rows:
        if r.get("is_service_fee"):
            service_fee = r
            continue
        sec = r.get("section", "other")
        rows_by_sec.setdefault(sec, []).append(r)

    # 6. Bölümleri yaz
    current_row = start_row
    subtotal_rows: dict[str, list[int]] = {}   # col_letter -> [subtotal satır numaraları]

    for sec in SECTIONS_ORDER:
        sec_rows = rows_by_sec.get(sec, [])
        if not sec_rows:
            continue

        sec_label = SECTION_LABELS.get(sec, sec)

        # Bölüm başlık satırı
        _safe_set(ws, current_row, label_col, sec_label,
                  font=_HDR_FONT(), fill=_HDR_FILL())
        current_row += 1

        sec_data_start = current_row

        # Veri satırları
        for row_data in sec_rows:
            for col_letter, field_name in col_defs.items():
                val = _row_value(field_name, row_data, budget, currency)
                _safe_set(ws, current_row, col_letter, val, font=_DAT_FONT())
            # Formül kolonları
            for col_letter, tpl in formula_cols.items():
                formula = tpl.replace("{row}", str(current_row))
                _safe_set(ws, current_row, col_letter, formula, font=_DAT_FONT())
            current_row += 1

        sec_data_end = current_row - 1
        print(f"[FILLER] sec={sec} hdr={sec_data_start-1} data={sec_data_start}..{sec_data_end} subtotal={current_row}", flush=True)

        # Ara toplam satırı
        sub_label = SECTION_SUBTOTAL_LABELS.get(sec, f"{sec_label} Ara Toplam")
        _safe_set(ws, current_row, label_col, sub_label,
                  font=_SUB_FONT(), fill=_SUB_FILL())
        for col_letter in formula_cols:
            formula = f"=SUM({col_letter}{sec_data_start}:{col_letter}{sec_data_end})"
            _safe_set(ws, current_row, col_letter, formula,
                      font=_SUB_FONT(), fill=_SUB_FILL())
            subtotal_rows.setdefault(col_letter, []).append(current_row)
        current_row += 1

    # Hizmet bedeli satırı
    if service_fee:
        sf_label = service_fee.get("service_name") or "Hizmet Bedeli"
        _safe_set(ws, current_row, label_col, sf_label, font=_DAT_FONT())
        for col_letter, field_name in col_defs.items():
            val = _row_value(field_name, service_fee, budget, currency)
            _safe_set(ws, current_row, col_letter, val, font=_DAT_FONT())
        for col_letter, tpl in formula_cols.items():
            formula = tpl.replace("{row}", str(current_row))
            _safe_set(ws, current_row, col_letter, formula, font=_DAT_FONT())
            subtotal_rows.setdefault(col_letter, []).append(current_row)
        current_row += 1

    # Genel toplam satırı
    if subtotal_rows:
        _safe_set(ws, current_row, label_col, "GENEL TOPLAM (KDV Hariç)",
                  font=_TOT_FONT(), fill=_TOT_FILL())
        for col_letter, rows in subtotal_rows.items():
            refs = "+".join(f"{col_letter}{r}" for r in rows)
            _safe_set(ws, current_row, col_letter, f"={refs}",
                      font=_TOT_FONT(), fill=_TOT_FILL())
        current_row += 1

    # 7. Kullanılmayan eski satırları sil
    if clear_end >= current_row:
        try:
            ws.delete_rows(current_row, clear_end - current_row + 1)
        except Exception:
            pass


# ── Tek bütçe export ──────────────────────────────────────────────────────────
def fill_customer_template(
    template_path: str,
    cell_map: dict,
    budget,
    request,
    customer,
    creator,
) -> io.BytesIO:
    """Müşteri template'ini tek bütçe ile doldurur."""
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl kurulu değil")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template bulunamadı: {template_path}")

    wb = openpyxl.load_workbook(template_path)
    sheet_name = (cell_map.get("data_block") or {}).get("sheet")
    ws = (wb[sheet_name]
          if sheet_name and sheet_name in wb.sheetnames
          else wb.active)

    _fill_ws(ws, cell_map, budget, request, customer, creator)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Çok bütçe export (her bütçe ayrı sheet) ──────────────────────────────────
def fill_customer_template_multi(
    template_path: str,
    cell_map: dict,
    entries: list,
) -> io.BytesIO:
    """
    Birden fazla bütçeyi tek dosyada, her biri template kopyası olarak doldurur.
    entries: [{"budget": ..., "request": ..., "customer": ..., "creator": ...}]
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl kurulu değil")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template bulunamadı: {template_path}")

    sheet_name = (cell_map.get("data_block") or {}).get("sheet")
    used_titles: list[str] = []

    wb = openpyxl.load_workbook(template_path)
    src_ws = (wb[sheet_name]
              if sheet_name and sheet_name in wb.sheetnames
              else wb.active)

    for i, entry in enumerate(entries):
        b   = entry["budget"]
        req = entry.get("request")
        cus = entry.get("customer")
        cre = entry.get("creator")

        ws = src_ws if i == 0 else wb.copy_worksheet(src_ws)

        # Sheet adı = mekan/otel adı
        raw_title = (b.venue_name or f"Bütçe {i+1}")[:28].strip()
        title = raw_title
        suffix = 2
        while title in used_titles:
            title = f"{raw_title[:25]} {suffix}"
            suffix += 1
        used_titles.append(title)
        ws.title = title

        _fill_ws(ws, cell_map, b, req, cus, cre)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
