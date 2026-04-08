"""
Claude API ile Excel template analizi ve cell_map öğrenmesi.

Kullanım:
    result = await analyze_template("/path/to/template.xlsx", api_key="sk-ant-...")
    if result["error"]:
        print(result["error"])
    else:
        cell_map = result["cell_map"]
        # → customer.excel_config_json'a kaydet
"""
from __future__ import annotations

import json
import os

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── Sistem promptu ─────────────────────────────────────────────────────────────
_SYSTEM_PROMPT = """Sen bir Excel template analiz uzmanısın.
Sana bir Excel dosyasının içeriği JSON formatında verilecek (satır × sütun matrisi).
Bu template bir etkinlik organizasyon şirketinin müşterisine ait RFQ / fiyat teklifi formatıdır.

Görevin: Template'deki hücreleri E-dem bütçe sisteminin alanlarıyla eşleştirerek cell_map JSON döndür.

──────────────────────────────────────
E-DEM ALANLARI
──────────────────────────────────────
HEADER alanları (tek bir hücreye yazılır):
  event_name      → etkinlik adı
  ref_no          → referans numarası (TOP-ABC-2504-001 gibi)
  check_in        → etkinlik başlangıç tarihi
  check_out       → etkinlik bitiş tarihi
  venue_name      → mekan adı
  customer_name   → müşteri / firma adı
  creator_name    → teklifi hazırlayan kişi adı
  eur_rate        → 1 EUR = ? TL kur değeri (sayı)
  usd_rate        → 1 USD = ? TL kur değeri (sayı)
  attendee_count  → katılımcı sayısı
  city            → şehir(ler)

SATIR alanları (her bütçe kalemi için tekrarlanan):
  service_name    → hizmet / kalem adı
  notes           → not / açıklama
  unit            → birim (Gece, Kişi, Adet, Gün...)
  qty             → miktar / kişi sayısı
  nights          → gece / gün sayısı
  sale_price      → birim satış fiyatı (KDV hariç, teklif para biriminde)
  sale_price_inc  → birim satış fiyatı (KDV dahil)
  vat_rate        → KDV yüzdesi (20, 10, 0 gibi)
  vat_pct         → KDV oranı (0.20, 0.10 gibi)
  total_excl      → toplam (KDV hariç) = sale_price × qty × nights
  total_incl      → toplam (KDV dahil)
  sale_price_eur  → birim satış fiyatı Euro cinsinden
  total_eur       → toplam Euro cinsinden
  sale_price_usd  → birim satış fiyatı USD cinsinden
  total_usd       → toplam USD cinsinden

──────────────────────────────────────
ÇIKTI FORMATI (sadece JSON, başka metin YAZMA)
──────────────────────────────────────
{
  "vat_mode": "exclusive",
  "header": {
    "B3": "event_name",
    "C4": "ref_no",
    "F4": "eur_rate"
  },
  "data_block": {
    "start_row": 9,
    "end_anchor_text": "ARA TOPLAM",
    "sheet": null,
    "columns": {
      "B": "service_name",
      "C": "notes",
      "D": "nights",
      "E": "qty",
      "F": "sale_price_eur",
      "H": "sale_price",
      "J": "total_incl"
    },
    "section_header_col": "B"
  }
}

KURALLAR:
1. data_block.start_row: İlk gerçek VERİ satırı numarası (başlık satırı değil)
2. end_anchor_text: Veri bloğunun bittiğini gösteren hücre içeriği (varsa, yoksa null)
3. section_header_col: Kategoriler ayrı satır olarak yazılıyorsa hangi sütun (yoksa null)
4. vat_mode: Template'de KDV hariç fiyat görünüyorsa "exclusive", KDV dahil ise "inclusive"
5. Emin olmadığın header alanlarını ekleme — boş {} daha iyi
6. SADECE geçerli bir JSON döndür — başka hiçbir metin yazma"""


def parse_template_structure(template_path: str, max_rows: int = 30) -> list[list]:
    """
    Excel template'ini JSON-serializable matrise dönüştürür.
    AI'ya göndermek için kullanılır.
    """
    if not OPENPYXL_OK:
        raise RuntimeError("openpyxl kurulu değil")

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    result = []
    for row in ws.iter_rows(min_row=1,
                            max_row=min(max_rows, ws.max_row),
                            values_only=True):
        serialized = []
        for cell in row:
            if cell is None:
                serialized.append(None)
            elif hasattr(cell, "isoformat"):
                serialized.append(str(cell))
            else:
                serialized.append(cell)
        if any(c is not None for c in serialized):
            result.append(serialized)

    return result


def _extract_json(text: str) -> str:
    """Yanıt metninden JSON bloğunu çıkarır."""
    text = text.strip()
    # ```json ... ``` bloğu varsa içini al
    if "```" in text:
        parts = text.split("```")
        for part in parts:
            part = part.strip()
            if part.startswith("json"):
                part = part[4:].strip()
            if part.startswith("{"):
                return part
    # Doğrudan JSON ise
    if text.startswith("{"):
        return text
    # İçinde JSON var mı?
    start = text.find("{")
    end   = text.rfind("}")
    if start != -1 and end != -1:
        return text[start:end + 1]
    return text


async def analyze_template(
    template_path: str,
    api_key: str | None = None,
    model: str = "claude-opus-4-6",
) -> dict:
    """
    Claude API kullanarak template yapısını analiz eder.

    Returns:
        {
            "cell_map": dict,       # başarılıysa dolu
            "raw_response": str,    # ham Claude yanıtı
            "error": str | None,    # hata mesajı
        }
    """
    try:
        import anthropic
    except ImportError:
        return {
            "cell_map": {},
            "raw_response": "",
            "error": "anthropic paketi kurulu değil. pip install anthropic",
        }

    key = api_key or os.environ.get("ANTHROPIC_API_KEY", "")
    if not key:
        return {
            "cell_map": {},
            "raw_response": "",
            "error": "ANTHROPIC_API_KEY ortam değişkeni bulunamadı",
        }

    try:
        structure = parse_template_structure(template_path, max_rows=30)
    except Exception as exc:
        return {"cell_map": {}, "raw_response": "", "error": str(exc)}

    user_msg = (
        "Excel template yapısı (satır listesi, her satır hücre değerlerini içerir):\n\n"
        f"```json\n{json.dumps(structure, ensure_ascii=False, indent=2)}\n```\n\n"
        "Bu template için E-dem cell_map JSON'ını döndür."
    )

    raw = ""
    try:
        client = anthropic.Anthropic(api_key=key)
        response = client.messages.create(
            model=model,
            max_tokens=1024,
            system=_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_msg}],
        )
        raw = response.content[0].text
        cell_map = json.loads(_extract_json(raw))
        return {"cell_map": cell_map, "raw_response": raw, "error": None}

    except json.JSONDecodeError as exc:
        return {
            "cell_map": {},
            "raw_response": raw,
            "error": f"Claude yanıtı JSON parse hatası: {exc}",
        }
    except Exception as exc:
        return {
            "cell_map": {},
            "raw_response": raw,
            "error": str(exc),
        }
