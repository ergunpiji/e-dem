"""
Haftalık Ödeme Listesi — Genel Müdür için
GET /payments/weekly                   — sayfa
POST /payments/weekly/decide            — onayla / reddet / ertele / yöntem değiştir
POST /payments/settings/weekday         — admin: ödeme günü ayarla
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from auth import get_current_user, require_admin
from database import get_db
from models import (
    User, Invoice, Cheque, CreditCardStatement, CreditCard, CreditCardTxn,
    Employee, SalaryPayment, PayrollDecision, SystemSetting,
    BankAccount, BankMovement, CashEntry,
    PAYMENT_METHODS,
)
from templates_config import templates


router = APIRouter(prefix="/payments", tags=["payments"])

WEEKDAYS_TR = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]
PAYMENT_METHOD_LABELS = dict(PAYMENT_METHODS)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _get_payment_weekday(db: Session) -> int:
    s = db.query(SystemSetting).filter(SystemSetting.key == "payment_weekday").first()
    if not s:
        return 2
    try:
        v = int(s.value)
        return v if 0 <= v <= 6 else 2
    except (TypeError, ValueError):
        return 2


def _set_payment_weekday(db: Session, weekday: int) -> None:
    s = db.query(SystemSetting).filter(SystemSetting.key == "payment_weekday").first()
    if s:
        s.value = str(weekday)
    else:
        db.add(SystemSetting(key="payment_weekday", value=str(weekday)))


def _next_payment_date(weekday: int, today: Optional[date] = None) -> date:
    if today is None:
        today = date.today()
    days_ahead = (weekday - today.weekday()) % 7
    return today + timedelta(days=days_ahead)


def _show_in_list(item, ref_date: date) -> bool:
    """Listede gösterilecek mi? Onaylananlar görünür kalır (yeşil ışık,
    fiili ödeme yapılınca status değişip filtreden düşer). Reddedilenler
    listeden kaldırılır. Ertelenenler ancak vade geldiğinde tekrar çıkar."""
    d = item.gm_decision
    if d is None:
        return True
    if d == "approved":
        return True
    if d == "rejected":
        return False
    if d == "postponed":
        return bool(item.gm_postpone_until and item.gm_postpone_until <= ref_date)
    return True


def _is_actionable(item, ref_date: date) -> bool:
    """Yeni karar verilebilir mi? (Onayla/Reddet butonu + checkbox görünür)
    - Karar verilmemiş kalemler: actionable
    - Ertelenmiş + vade geldi: actionable (yeniden karar)
    - Kısmi onay + vade geldi: actionable (kalan için karar)
    - Tam onay veya henüz vadeye gelmemiş erteleme: not actionable
    """
    d = item.gm_decision
    if d is None:
        return True
    if d == "rejected":
        return False
    pp = getattr(item, "gm_postpone_until", None)
    if d == "postponed":
        return bool(pp and pp <= ref_date)
    if d == "approved":
        # Kısmi onaylı (gm_approved_amount set) ve vade geldi → tekrar karar zamanı
        approved_amt = getattr(item, "gm_approved_amount", None)
        if approved_amt and pp and pp <= ref_date:
            return True
        return False
    return False


def _cash_total(db) -> float:
    ins = db.query(func.sum(CashEntry.amount)).filter(CashEntry.entry_type == "giris").scalar() or 0
    outs = db.query(func.sum(CashEntry.amount)).filter(CashEntry.entry_type == "cikis").scalar() or 0
    return ins - outs


def _bank_total(db) -> float:
    accounts = db.query(BankAccount).all()
    total = 0.0
    for a in accounts:
        opening = a.opening_balance or 0
        ins = db.query(func.sum(BankMovement.amount)).filter(
            BankMovement.account_id == a.id, BankMovement.movement_type == "giris"
        ).scalar() or 0
        outs = db.query(func.sum(BankMovement.amount)).filter(
            BankMovement.account_id == a.id, BankMovement.movement_type == "cikis"
        ).scalar() or 0
        total += opening + ins - outs
    return total


def _cc_outstanding(db, card_id) -> float:
    unpaid = db.query(func.sum(CreditCardStatement.total_amount)).filter(
        CreditCardStatement.card_id == card_id,
        CreditCardStatement.status == "unpaid",
    ).scalar() or 0
    unassigned = db.query(func.sum(CreditCardTxn.amount)).filter(
        CreditCardTxn.card_id == card_id,
        CreditCardTxn.statement_id == None,  # noqa: E711
        CreditCardTxn.is_refund == False,  # noqa: E712
    ).scalar() or 0
    return unpaid + unassigned


def _payroll_due(db, period: str) -> dict:
    employees = db.query(Employee).filter(Employee.active == True).all()  # noqa: E712
    paid_ids = {p.employee_id for p in db.query(SalaryPayment).filter(SalaryPayment.period == period).all()}
    unpaid = [e for e in employees if e.id not in paid_ids]
    total = sum(e.net_salary or 0 for e in unpaid)
    return {"unpaid_count": len(unpaid), "total": total, "employees": unpaid}


def _method_label(code: Optional[str]) -> str:
    if not code:
        return "—"
    return PAYMENT_METHOD_LABELS.get(code, code)


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.get("/weekly", response_class=HTMLResponse, name="weekly_payments")
async def weekly_payments_view(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    weekday = _get_payment_weekday(db)
    next_date = _next_payment_date(weekday)
    next_payment_after = next_date + timedelta(days=7)  # bir sonraki ödeme günü
    today = date.today()
    period = today.strftime("%Y-%m")

    # Faturalar
    inv_q = db.query(Invoice).filter(
        Invoice.invoice_type == "gelen",
        Invoice.status.in_(["approved", "partial"]),
        Invoice.due_date != None,  # noqa: E711
        Invoice.due_date <= next_date,
    ).order_by(Invoice.due_date.asc()).all()
    invoices = [i for i in inv_q if i.remaining > 0 and _show_in_list(i, next_date)]

    # Çekler
    chq_q = db.query(Cheque).filter(
        Cheque.cheque_type == "verilen",
        Cheque.status == "beklemede",
        Cheque.due_date <= next_date,
    ).order_by(Cheque.due_date.asc()).all()
    cheques = [c for c in chq_q if _show_in_list(c, next_date)]

    # KK Ekstreleri (Çarşamba kuralından bağımsız — kendi vadesi geçerli)
    cc_q = db.query(CreditCardStatement).filter(
        CreditCardStatement.status == "unpaid",
        CreditCardStatement.due_date <= next_date,
    ).order_by(CreditCardStatement.due_date.asc()).all()
    cc_stmts = [s for s in cc_q if _show_in_list(s, next_date)]

    # Maaş — bu ay ödenmemiş aktifler
    payroll_info = _payroll_due(db, period)
    payroll_decision = db.query(PayrollDecision).filter(PayrollDecision.period == period).first()
    payroll_show = (
        payroll_info["total"] > 0 and (
            payroll_decision is None
            or _show_in_list(payroll_decision, next_date)
        )
    )

    # Birleşik liste — template tek tablo render eder
    items = []
    for inv in invoices:
        items.append({
            "kalem_type": "invoice",
            "kalem_id": inv.id,
            "type_label": "Fatura",
            "type_color": "primary",
            "party": inv.vendor.name if inv.vendor else "—",
            "party_url": f"/vendors/{inv.vendor_id}" if inv.vendor_id else None,
            "ref_no": inv.reference.ref_no if inv.reference else "—",
            "ref_url": f"/references/{inv.ref_id}" if inv.ref_id else None,
            "detail_url": f"/invoices/{inv.id}",
            "due_date": inv.due_date,
            "amount": inv.remaining,
            "method_code": inv.gm_method_override or inv.payment_method,
            "method_label": _method_label(inv.gm_method_override or inv.payment_method),
            "method_override": inv.gm_method_override,
            "gm_decision": inv.gm_decision,
            "gm_postpone_until": inv.gm_postpone_until,
            "actionable": _is_actionable(inv, next_date),
            "gm_decision_note": inv.gm_decision_note,
            "gm_approved_amount": inv.gm_approved_amount,
            "preparer_note": inv.preparer_note,
        })

    for c in cheques:
        party = (c.vendor.name if c.vendor else None) or "—"
        items.append({
            "kalem_type": "cheque",
            "kalem_id": c.id,
            "type_label": "Çek",
            "type_color": "warning",
            "party": party,
            "ref_no": c.cheque_no or "—",
            "ref_url": None,
            "detail_url": "/cheques",
            "due_date": c.due_date,
            "amount": c.amount,
            "method_code": c.gm_method_override or "cek",
            "method_label": _method_label(c.gm_method_override or "cek"),
            "method_override": c.gm_method_override,
            "gm_decision": c.gm_decision,
            "gm_postpone_until": c.gm_postpone_until,
            "actionable": _is_actionable(c, next_date),
            "gm_decision_note": c.gm_decision_note,
            "gm_approved_amount": c.gm_approved_amount,
            "preparer_note": c.preparer_note,
        })

    for s in cc_stmts:
        items.append({
            "kalem_type": "cc_statement",
            "kalem_id": s.id,
            "type_label": "KK Ekstre",
            "type_color": "danger",
            "party": s.card.name if s.card else "—",
            "party_url": f"/credit-cards/{s.card_id}" if s.card_id else None,
            "ref_no": "—",
            "ref_url": None,
            "detail_url": f"/credit-cards/{s.card_id}",
            "due_date": s.due_date,
            "amount": s.total_amount,
            "method_code": s.gm_method_override or "kredi_karti",
            "method_label": _method_label(s.gm_method_override or "kredi_karti"),
            "method_override": s.gm_method_override,
            "gm_decision": s.gm_decision,
            "gm_postpone_until": s.gm_postpone_until,
            "actionable": _is_actionable(s, next_date),
            "gm_decision_note": s.gm_decision_note,
            "gm_approved_amount": s.gm_approved_amount,
            "preparer_note": s.preparer_note,
        })

    if payroll_show:
        method = (payroll_decision.gm_method_override if payroll_decision else None) or "banka"
        items.append({
            "kalem_type": "payroll",
            "kalem_id": period,
            "type_label": "Maaş",
            "type_color": "info",
            "party": f"Personel ({payroll_info['unpaid_count']} kişi)",
            "ref_no": period,
            "ref_url": None,
            "detail_url": "/employees",
            "due_date": None,
            "amount": payroll_info["total"],
            "method_code": method,
            "method_label": _method_label(method),
            "method_override": payroll_decision.gm_method_override if payroll_decision else None,
            "gm_decision": payroll_decision.gm_decision if payroll_decision else None,
            "gm_postpone_until": payroll_decision.gm_postpone_until if payroll_decision else None,
            "actionable": _is_actionable(payroll_decision, next_date) if payroll_decision else True,
            "gm_decision_note": payroll_decision.gm_decision_note if payroll_decision else None,
            "gm_approved_amount": payroll_decision.gm_approved_amount if payroll_decision else None,
            "preparer_note": payroll_decision.preparer_note if payroll_decision else None,
            "party_url": "/employees",
        })

    # Vade tarihine göre sırala (maaş sona)
    items.sort(key=lambda x: (x["due_date"] is None, x["due_date"] or date.max))

    grand_total = sum(it["amount"] for it in items)

    # Özet
    cash_balance = _cash_total(db)
    bank_balance = _bank_total(db)
    cards = db.query(CreditCard).order_by(CreditCard.name).all()
    card_summary = []
    for c in cards:
        used = _cc_outstanding(db, c.id)
        # Sonraki ödenmemiş ekstrenin vadesi; yoksa kartın statement_day + payment_offset
        # ayarlarından bir sonraki ödeme tarihini hesapla.
        next_stmt = db.query(CreditCardStatement).filter(
            CreditCardStatement.card_id == c.id,
            CreditCardStatement.status == "unpaid",
        ).order_by(CreditCardStatement.due_date.asc()).first()
        if next_stmt:
            next_due = next_stmt.due_date
        elif c.statement_day:
            sd = c.statement_day
            offset = c.payment_offset_days or 10
            t = date.today()
            stmt_month = t.month if t.day <= sd else (t.month % 12 + 1)
            stmt_year = t.year + (1 if (t.day > sd and t.month == 12) else 0)
            try:
                stmt_d = date(stmt_year, stmt_month, sd)
            except ValueError:
                import calendar as _cal
                stmt_d = date(stmt_year, stmt_month, _cal.monthrange(stmt_year, stmt_month)[1])
            next_due = stmt_d + timedelta(days=offset)
        else:
            next_due = None
        card_summary.append({
            "id": c.id, "name": c.name,
            "limit": c.credit_limit or 0,
            "used": used,
            "available": (c.credit_limit or 0) - used,
            "next_due": next_due,
        })
    cc_total_used = sum(c["used"] for c in card_summary)
    cc_total_available = sum(c["available"] for c in card_summary)

    return templates.TemplateResponse(
        "payments/weekly.html",
        {
            "request": request, "current_user": current_user,
            "page_title": "Haftalık Ödeme Listesi",
            "next_date": next_date,
            "next_payment_after": next_payment_after,
            "next_payment_after_iso": next_payment_after.isoformat(),
            "weekday": weekday,
            "weekday_name": WEEKDAYS_TR[weekday],
            "weekdays_tr": WEEKDAYS_TR,
            "items": items,
            "grand_total": grand_total,
            "cash_balance": cash_balance,
            "bank_balance": bank_balance,
            "card_summary": card_summary,
            "cc_total_used": cc_total_used,
            "cc_total_available": cc_total_available,
            "payment_methods": PAYMENT_METHODS,
        },
    )


@router.post("/weekly/decide", name="weekly_payment_decide")
async def weekly_payment_decide(
    kalem_type: str = Form(...),
    kalem_id: str = Form(...),
    action: str = Form(...),
    postpone_date: str = Form(""),
    method_override: str = Form(""),
    note: str = Form(""),
    approved_amount: str = Form(""),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if action not in ("approve", "reject", "postpone", "method"):
        raise HTTPException(400, "Geçersiz aksiyon")
    # Onay/Red sadece Genel Müdür yetkisinde; ertele/yöntem değişikliği herkese açık
    if action in ("approve", "reject") and not current_user.is_approver:
        raise HTTPException(403, "Onay/Red için Genel Müdür yetkisi gereklidir.")

    if kalem_type == "invoice":
        item = db.query(Invoice).get(int(kalem_id))
    elif kalem_type == "cheque":
        item = db.query(Cheque).get(int(kalem_id))
    elif kalem_type == "cc_statement":
        item = db.query(CreditCardStatement).get(int(kalem_id))
    elif kalem_type == "payroll":
        item = db.query(PayrollDecision).filter(PayrollDecision.period == kalem_id).first()
        if not item:
            item = PayrollDecision(period=kalem_id)
            db.add(item)
            db.flush()
    else:
        raise HTTPException(400, "Geçersiz kalem tipi")

    if not item:
        raise HTTPException(404, "Kalem bulunamadı")

    now = datetime.utcnow()
    note_clean = (note or "").strip() or None
    if action == "approve":
        # Kalemin tam tutarı (kısmi onayda kıyas için)
        if kalem_type == "invoice":
            full_amount = item.remaining
        elif kalem_type == "cheque":
            full_amount = item.amount
        elif kalem_type == "cc_statement":
            full_amount = item.total_amount
        else:  # payroll
            full_amount = _payroll_due(db, item.period)["total"]

        # approved_amount opsiyonel: boş veya tam tutar → tam onay; daha az → kısmi
        try:
            req_amt = float(approved_amount) if (approved_amount or "").strip() else full_amount
        except (ValueError, TypeError) as exc:
            raise HTTPException(400, "Geçersiz onay tutarı") from exc
        if req_amt <= 0 or req_amt > round(full_amount + 0.01, 2):
            raise HTTPException(400, "Onay tutarı 0 ile tam tutar arasında olmalı")

        is_partial = req_amt < round(full_amount - 0.01, 2)
        item.gm_decision = "approved"
        item.gm_decision_at = now
        item.gm_decision_by = current_user.id
        item.gm_decision_note = note_clean
        if is_partial:
            # Kısmi onay: kalan kısmı için erteleme tarihi gerekli
            try:
                pd = date.fromisoformat(postpone_date)
            except (ValueError, TypeError) as exc:
                raise HTTPException(400, "Kısmi onay için erteleme tarihi gerekli") from exc
            if pd <= date.today():
                raise HTTPException(400, "Erteleme tarihi gelecekte olmalı")
            item.gm_approved_amount = round(req_amt, 2)
            item.gm_postpone_until = pd
        else:
            item.gm_approved_amount = None
            item.gm_postpone_until = None
    elif action == "reject":
        item.gm_decision = "rejected"
        item.gm_decision_at = now
        item.gm_decision_by = current_user.id
        item.gm_postpone_until = None
        item.gm_decision_note = note_clean
    elif action == "postpone":
        try:
            new_date = date.fromisoformat(postpone_date)
        except (ValueError, TypeError):
            raise HTTPException(400, "Geçersiz erteleme tarihi")
        if new_date <= date.today():
            raise HTTPException(400, "Erteleme tarihi gelecekte olmalı")
        item.gm_decision = "postponed"
        item.gm_decision_at = now
        item.gm_decision_by = current_user.id
        item.gm_postpone_until = new_date
    elif action == "method":
        valid = {m[0] for m in PAYMENT_METHODS}
        if method_override not in valid:
            raise HTTPException(400, "Geçersiz ödeme yöntemi")
        item.gm_method_override = method_override

    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)


@router.get("/weekly/export", name="weekly_payment_export")
async def weekly_payment_export(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Haftalık ödeme listesini Excel olarak indir."""
    weekday = _get_payment_weekday(db)
    next_date = _next_payment_date(weekday)
    today = date.today()
    period = today.strftime("%Y-%m")

    # View ile aynı filtreleme — yeniden hesapla
    invoices = db.query(Invoice).filter(
        Invoice.invoice_type == "gelen",
        Invoice.status.in_(["approved", "partial"]),
        Invoice.due_date.is_not(None),
        Invoice.due_date <= next_date,
    ).order_by(Invoice.due_date.asc()).all()
    invoices = [i for i in invoices if i.remaining > 0 and _show_in_list(i, next_date)]

    cheques = db.query(Cheque).filter(
        Cheque.cheque_type == "verilen",
        Cheque.status == "beklemede",
        Cheque.due_date <= next_date,
    ).order_by(Cheque.due_date.asc()).all()
    cheques = [c for c in cheques if _show_in_list(c, next_date)]

    cc_stmts = db.query(CreditCardStatement).filter(
        CreditCardStatement.status == "unpaid",
        CreditCardStatement.due_date <= next_date,
    ).order_by(CreditCardStatement.due_date.asc()).all()
    cc_stmts = [s for s in cc_stmts if _show_in_list(s, next_date)]

    payroll_info = _payroll_due(db, period)
    payroll_decision = db.query(PayrollDecision).filter(
        PayrollDecision.period == period
    ).first()
    payroll_show = (
        payroll_info["total"] > 0 and (
            payroll_decision is None or _show_in_list(payroll_decision, next_date)
        )
    )

    decision_label = {
        "approved": "Onaylandı",
        "rejected": "Reddedildi",
        "postponed": "Ertelendi",
    }

    def _decision_text(d):
        return decision_label.get(d, "Bekliyor")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Odeme Listesi"

    headers = ["Tip", "Tedarikci/Ilgili", "Referans", "Vade", "Tutar (TL)",
               "Yontem", "Durum", "Onaylanan (TL)", "Not"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = []
    for inv in invoices:
        rows.append([
            "Fatura",
            inv.vendor.name if inv.vendor else "—",
            inv.reference.ref_no if inv.reference else "—",
            inv.due_date.strftime("%d.%m.%Y") if inv.due_date else "",
            float(inv.remaining or 0),
            _method_label(inv.gm_method_override or inv.payment_method),
            _decision_text(inv.gm_decision),
            float(inv.gm_approved_amount) if inv.gm_approved_amount else None,
            inv.gm_decision_note or "",
        ])
    for c in cheques:
        rows.append([
            "Cek",
            (c.vendor.name if c.vendor else "") or "—",
            c.cheque_no or "—",
            c.due_date.strftime("%d.%m.%Y") if c.due_date else "",
            float(c.amount or 0),
            _method_label(c.gm_method_override or "cek"),
            _decision_text(c.gm_decision),
            float(c.gm_approved_amount) if c.gm_approved_amount else None,
            c.gm_decision_note or "",
        ])
    for s in cc_stmts:
        rows.append([
            "KK Ekstre",
            s.card.name if s.card else "—",
            "—",
            s.due_date.strftime("%d.%m.%Y") if s.due_date else "",
            float(s.total_amount or 0),
            _method_label(s.gm_method_override or "kredi_karti"),
            _decision_text(s.gm_decision),
            float(s.gm_approved_amount) if s.gm_approved_amount else None,
            s.gm_decision_note or "",
        ])
    if payroll_show:
        method = (payroll_decision.gm_method_override
                  if payroll_decision else None) or "banka"
        approved_amt = (payroll_decision.gm_approved_amount
                        if payroll_decision and payroll_decision.gm_approved_amount
                        else None)
        note_val = (payroll_decision.gm_decision_note
                    if payroll_decision else None) or ""
        rows.append([
            "Maas",
            f"Personel ({payroll_info['unpaid_count']} kisi)",
            period, "",
            float(payroll_info["total"] or 0),
            _method_label(method),
            _decision_text(payroll_decision.gm_decision if payroll_decision else None),
            float(approved_amt) if approved_amt else None,
            note_val,
        ])

    grand_total = 0.0
    for r in rows:
        ws.append(r)
        grand_total += r[4] or 0

    # Toplam satırı
    ws.append([])
    total_row = ws.max_row + 1
    label_cell = ws.cell(row=total_row, column=4, value="GENEL TOPLAM")
    label_cell.font = Font(bold=True)
    label_cell.alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=5, value=grand_total)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'

    # Sütun genişlikleri (sabit harflerle)
    widths = {"A": 12, "B": 38, "C": 18, "D": 12, "E": 16,
              "F": 18, "G": 14, "H": 16, "I": 30}
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    # Para sütunlarında (E, H) sayı formatı
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in (5, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"odeme-listesi-{next_date.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/weekly/bulk-approve", name="weekly_payment_bulk_approve")
async def weekly_payment_bulk_approve(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not current_user.is_approver:
        raise HTTPException(403, "Toplu onay için Genel Müdür yetkisi gereklidir.")
    form = await request.form()
    entries = form.getlist("items")
    note_clean = (form.get("note", "") or "").strip() or None
    now = datetime.utcnow()
    count = 0
    for entry in entries:
        if ":" not in entry:
            continue
        kalem_type, kalem_id = entry.split(":", 1)
        if kalem_type == "invoice":
            item = db.query(Invoice).get(int(kalem_id))
        elif kalem_type == "cheque":
            item = db.query(Cheque).get(int(kalem_id))
        elif kalem_type == "cc_statement":
            item = db.query(CreditCardStatement).get(int(kalem_id))
        elif kalem_type == "payroll":
            item = db.query(PayrollDecision).filter(PayrollDecision.period == kalem_id).first()
            if not item:
                item = PayrollDecision(period=kalem_id)
                db.add(item)
                db.flush()
        else:
            continue
        if not item:
            continue
        item.gm_decision = "approved"
        item.gm_decision_at = now
        item.gm_decision_by = current_user.id
        item.gm_postpone_until = None
        if note_clean:
            item.gm_decision_note = note_clean
        count += 1
    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)


@router.post("/weekly/preparer-note", name="weekly_payment_preparer_note")
async def weekly_payment_preparer_note(
    kalem_type: str = Form(...),
    kalem_id: str = Form(...),
    note: str = Form(""),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Listeyi hazırlayan kişinin GM'e yönelik notu — herkes kaydedebilir."""
    note_clean = (note or "").strip() or None
    if kalem_type == "invoice":
        item = db.query(Invoice).get(int(kalem_id))
    elif kalem_type == "cheque":
        item = db.query(Cheque).get(int(kalem_id))
    elif kalem_type == "cc_statement":
        item = db.query(CreditCardStatement).get(int(kalem_id))
    elif kalem_type == "payroll":
        item = db.query(PayrollDecision).filter(PayrollDecision.period == kalem_id).first()
        if not item:
            item = PayrollDecision(period=kalem_id)
            db.add(item)
            db.flush()
    else:
        raise HTTPException(400, "Geçersiz kalem tipi")
    if not item:
        raise HTTPException(404, "Kalem bulunamadı")
    item.preparer_note = note_clean
    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)


@router.post("/settings/weekday", name="weekly_payment_set_weekday")
async def set_payment_weekday(
    weekday: int = Form(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if not (0 <= weekday <= 6):
        raise HTTPException(400, "Geçersiz gün (0-6 olmalı)")
    _set_payment_weekday(db, weekday)
    db.commit()
    return RedirectResponse(url="/payments/weekly", status_code=303)
