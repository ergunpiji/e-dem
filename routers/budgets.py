"""
E-dem — Bütçe yönetimi router'ı

Workflow:
  E-dem:   oluştur (draft_edem) → manager'a gönder (pending_manager) → revizyonu düzelt
  Manager: satış fiyatı gir (draft_manager) → onayla (approved) / revizyon iste / iptal et
  Admin:   her şeyi görür

Endpoints:
  GET    /budgets                    → liste (role-based)
  GET    /budgets/new                → E-dem: yeni bütçe formu
  POST   /budgets/new                → E-dem: oluştur
  GET    /budgets/{id}               → detay (role-based)
  GET    /budgets/{id}/edit          → E-dem: maliyet düzenle (draft_edem veya revision_requested)
  POST   /budgets/{id}/edit          → E-dem: kaydet
  POST   /budgets/{id}/send-to-manager → E-dem: manager'a gönder
  GET    /budgets/{id}/price         → Manager: satış fiyatı editörü
  POST   /budgets/{id}/price         → Manager: satış fiyatlarını kaydet (draft_manager)
  POST   /budgets/{id}/approve       → Manager: onayla → approved
  POST   /budgets/{id}/request-revision → Manager: revizyon iste
  POST   /budgets/{id}/cancel        → Manager: iptal et
  GET    /budgets/{id}/export        → Manager: Excel export (customer template kullan)
  POST   /budgets/{id}/delete        → E-dem/Admin: sil (sadece draft_edem)
"""

import io
import json
import os

from fastapi import APIRouter, Depends, Form, HTTPException, Request, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session

from auth import get_current_user
from database import get_db
from models import (
    Budget, Customer, Request as ReqModel, Service, SERVICE_CATEGORIES, User, _uuid, _now,
)

router = APIRouter(prefix="/budgets", tags=["budgets"])
from templates_config import templates

BUDGET_STATUS_LABELS = {
    "draft_edem":         "Taslak (E-dem)",
    "pending_manager":    "Manager Onayında",
    "draft_manager":      "Manager Düzenliyor",
    "approved":           "Onaylandı",
    "revision_requested": "Revizyon Bekleniyor",
    "cancelled":          "İptal Edildi",
}
BUDGET_STATUS_COLORS = {
    "draft_edem":         "secondary",
    "pending_manager":    "warning",
    "draft_manager":      "info",
    "approved":           "success",
    "revision_requested": "danger",
    "cancelled":          "dark",
}


def _can_edem_edit(budget: Budget) -> bool:
    return budget.budget_status in ("draft_edem", "revision_requested")


def _can_manager_price(budget: Budget) -> bool:
    return budget.budget_status not in ("approved", "cancelled")


@router.get("", response_class=HTMLResponse, name="budgets_list")
async def budgets_list(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(Budget)
    if current_user.role == "e_dem":
        query = query.filter(Budget.created_by == current_user.id)
    elif current_user.role == "project_manager":
        my_req_ids = [
            r.id for r in db.query(ReqModel)
            .filter(ReqModel.created_by == current_user.id)
            .all()
        ]
        query = query.filter(Budget.request_id.in_(my_req_ids))

    budgets = query.order_by(Budget.created_at.desc()).all()
    return templates.TemplateResponse("budgets/list.html", {
        "request":       request,
        "current_user":  current_user,
        "budgets":       budgets,
        "page_title":    "Bütçe Yönetimi" if current_user.role == "e_dem" else "Bütçeler",
        "status_labels": BUDGET_STATUS_LABELS,
        "status_colors": BUDGET_STATUS_COLORS,
    })


def _can_create_budget(user: User) -> bool:
    return user.role in ("admin", "e_dem", "project_manager")


SECTION_ORDER = ["accommodation", "meeting", "fb", "teknik", "dekor", "transfer", "tasarim", "other"]


def _default_vat(section: str) -> int:
    return 10 if section == "accommodation" else 20


def _calc_nights(date_from: str, date_to: str) -> int:
    """İki tarih arasındaki gece/gün sayısını hesapla (minimum 1)"""
    try:
        from datetime import date as dt
        d1 = dt.fromisoformat(str(date_from))
        d2 = dt.fromisoformat(str(date_to))
        n = (d2 - d1).days
        return max(1, n)
    except Exception:
        return 1


def _items_to_budget_rows(items: dict, req) -> list:
    """Talep items_json dict → bütçe satırı flat list"""
    rows = []
    for section in SECTION_ORDER:
        for item in items.get(section, []):
            date_from = str(item.get("date_from") or "")
            date_to   = str(item.get("date_to")   or "")
            # Tarih yoksa talepten al
            if section == "accommodation":
                date_from = date_from or str(req.accom_check_in or req.check_in or "")
                date_to   = date_to   or str(req.accom_check_out or req.check_out or "")
            elif section == "meeting":
                date_from = date_from or str(req.check_in  or "")
                date_to   = date_to   or str(req.check_out or "")

            nights = _calc_nights(date_from, date_to) if date_from and date_to else 1

            row = {
                "section":      section,
                "service_name": item.get("description", ""),
                "unit":         item.get("unit", "Adet"),
                "qty":          float(item.get("qty") or 1),
                "nights":       nights,
                "cost_price":   0,
                "sale_price":   0,
                "vat_rate":     _default_vat(section),
                "service_id":   item.get("service_id"),
                "quotes":       [],
            }
            if section == "accommodation":
                row["accom_in"]  = date_from
                row["accom_out"] = date_to
            elif section == "meeting":
                row["meeting_in"]  = date_from
                row["meeting_out"] = date_to
            rows.append(row)
    return rows


@router.get("/new", response_class=HTMLResponse, name="budgets_new")
async def budgets_new(
    request: Request,
    req_id: str = "",
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _can_create_budget(current_user):
        raise HTTPException(403)
    req = db.query(ReqModel).filter(ReqModel.id == req_id).first() if req_id else None
    services = db.query(Service).filter(Service.active == True).order_by(Service.category, Service.name).all()
    grouped_services: dict = {}
    for svc in services:
        grouped_services.setdefault(svc.category, []).append(svc.to_dict())

    # Talep kalemlerini bütçe satırlarına dönüştür
    initial_rows: list = []
    preferred_venues: list = []
    if req:
        items = req.items or {}
        if any(items.get(s) for s in SECTION_ORDER):
            initial_rows = _items_to_budget_rows(items, req)
        # Tercih edilen mekanları getir (mekan adı seçimi için)
        if req.preferred_venues:
            from models import Venue as VenueModel
            preferred_venues = (
                db.query(VenueModel)
                  .filter(VenueModel.id.in_(req.preferred_venues))
                  .order_by(VenueModel.name)
                  .all()
            )

    return templates.TemplateResponse("budgets/editor.html", {
        "request":            request,
        "current_user":       current_user,
        "budget":             None,
        "req":                req,
        "page_title":         "Yeni Bütçe",
        "service_categories": SERVICE_CATEGORIES,
        "grouped_services":   json.dumps(grouped_services, ensure_ascii=False),
        "initial_rows_json":  json.dumps(initial_rows, ensure_ascii=False),
        "preferred_venues":   preferred_venues,
    })


@router.post("/new", name="budgets_create")
async def budgets_create(
    request: Request,
    req_id:     str = Form(...),
    venue_name: str = Form(""),
    rows_json:  str = Form("[]"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _can_create_budget(current_user):
        raise HTTPException(403)

    # Satış fiyatlarını sıfırla (E-dem sadece maliyet girer)
    try:
        rows = json.loads(rows_json)
        for row in rows:
            if not row.get("is_service_fee"):
                row["sale_price"] = 0
        rows_json = json.dumps(rows, ensure_ascii=False)
    except Exception:
        pass

    budget = Budget(
        id=_uuid(),
        request_id=req_id,
        venue_name=venue_name.strip(),
        rows_json=rows_json,
        budget_status="draft_edem",
        created_by=current_user.id,
        created_at=_now(),
        updated_at=_now(),
    )
    db.add(budget)
    req = db.query(ReqModel).filter(ReqModel.id == req_id).first()
    if req and req.status in ("in_progress", "venues_contacted"):
        req.status = "budget_ready"
        req.updated_at = _now()
    db.commit()
    return RedirectResponse(url=f"/budgets/{budget.id}", status_code=status.HTTP_302_FOUND)


@router.get("/{budget_id}", response_class=HTMLResponse, name="budgets_detail")
async def budgets_detail(
    budget_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if not budget:
        return RedirectResponse(url="/budgets", status_code=status.HTTP_302_FOUND)
    req = db.query(ReqModel).filter(ReqModel.id == budget.request_id).first()
    rows_by_section: dict = {}
    for row in budget.rows:
        sec = row.get("section", "other")
        rows_by_section.setdefault(sec, []).append(row)

    # KDV rate bazlı döküm (detail sayfası alt özet için)
    vat_by_rate: dict = {}
    for row in budget.rows:
        qty    = float(row.get("qty", 1) or 1)
        nights = float(row.get("nights", 1) or 1)
        sale   = float(row.get("sale_price", 0) or 0)
        vrate  = float(row.get("vat_rate", 0) or 0)
        if row.get("is_service_fee"):
            qty, nights = 1, 1
        sale_sub = sale * qty * nights
        vat_amt  = round(sale_sub * (vrate / 100), 2)
        if vat_amt > 0 and vrate > 0:
            vat_by_rate[int(vrate)] = round(vat_by_rate.get(int(vrate), 0) + vat_amt, 2)
    vat_by_rate_sorted = sorted(vat_by_rate.items())  # [(10, 30000), (20, 5000)]

    return templates.TemplateResponse("budgets/detail.html", {
        "request":            request,
        "current_user":       current_user,
        "budget":             budget,
        "req":                req,
        "page_title":         f"Bütçe — {budget.venue_name or 'Yeni'}",
        "rows_by_section":    rows_by_section,
        "vat_by_rate":        vat_by_rate_sorted,
        "service_categories": SERVICE_CATEGORIES,
        "can_edem_edit":      _can_edem_edit(budget) and current_user.role in ("admin", "e_dem"),
        "can_manager_price":  _can_manager_price(budget) and current_user.role in ("admin", "project_manager"),
        "status_label":       BUDGET_STATUS_LABELS.get(budget.budget_status, budget.budget_status),
        "status_color":       BUDGET_STATUS_COLORS.get(budget.budget_status, "secondary"),
    })


@router.get("/{budget_id}/edit", response_class=HTMLResponse, name="budgets_edit")
async def budgets_edit(
    budget_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _can_create_budget(current_user):
        raise HTTPException(403)
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if not budget:
        return RedirectResponse(url="/budgets", status_code=status.HTTP_302_FOUND)
    if not _can_edem_edit(budget) and current_user.role != "admin":
        return RedirectResponse(url=f"/budgets/{budget_id}", status_code=status.HTTP_302_FOUND)
    req = db.query(ReqModel).filter(ReqModel.id == budget.request_id).first()
    services = db.query(Service).filter(Service.active == True).order_by(Service.category, Service.name).all()
    grouped_services: dict = {}
    for svc in services:
        grouped_services.setdefault(svc.category, []).append(svc.to_dict())
    return templates.TemplateResponse("budgets/editor.html", {
        "request":            request,
        "current_user":       current_user,
        "budget":             budget,
        "req":                req,
        "page_title":         f"Bütçe Düzenle — {budget.venue_name}",
        "service_categories": SERVICE_CATEGORIES,
        "grouped_services":   json.dumps(grouped_services, ensure_ascii=False),
        "initial_rows_json":  "[]",
        "preferred_venues":   [],
    })


@router.post("/{budget_id}/edit", name="budgets_update")
async def budgets_update(
    budget_id: str,
    venue_name: str = Form(""),
    rows_json:  str = Form("[]"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not _can_create_budget(current_user):
        raise HTTPException(403)
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if not budget:
        return RedirectResponse(url="/budgets", status_code=status.HTTP_302_FOUND)

    # E-dem satış fiyatı giremez — mevcut sale_price değerlerini sıfırla
    try:
        new_rows = json.loads(rows_json)
        for row in new_rows:
            if not row.get("is_service_fee"):
                row["sale_price"] = 0
        rows_json = json.dumps(new_rows, ensure_ascii=False)
    except Exception:
        pass

    budget.venue_name = venue_name.strip()
    budget.rows_json  = rows_json
    budget.updated_at = _now()
    db.commit()
    return RedirectResponse(url=f"/budgets/{budget_id}", status_code=status.HTTP_302_FOUND)


@router.post("/{budget_id}/send-to-manager", name="budgets_send_to_manager")
async def budgets_send_to_manager(
    budget_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("admin", "e_dem"):
        raise HTTPException(403)
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if budget and budget.budget_status in ("draft_edem", "revision_requested"):
        budget.budget_status = "pending_manager"
        budget.updated_at    = _now()
        db.commit()
    return RedirectResponse(url=f"/budgets/{budget_id}", status_code=status.HTTP_302_FOUND)


@router.get("/{budget_id}/price", response_class=HTMLResponse, name="budgets_price")
async def budgets_price(
    budget_id: str,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("admin", "project_manager"):
        raise HTTPException(403)
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if not budget:
        return RedirectResponse(url="/budgets", status_code=status.HTTP_302_FOUND)
    req = db.query(ReqModel).filter(ReqModel.id == budget.request_id).first()
    customer = db.query(Customer).filter(Customer.id == req.customer_id).first() if req and req.customer_id else None
    rows_by_section: dict = {}
    for row in budget.rows:
        sec = row.get("section", "other")
        rows_by_section.setdefault(sec, []).append(row)
    services = db.query(Service).filter(Service.active == True).order_by(Service.category, Service.name).all()
    grouped_services: dict = {}
    for svc in services:
        grouped_services.setdefault(svc.category, []).append(svc.to_dict())
    return templates.TemplateResponse("budgets/manager_editor.html", {
        "request":         request,
        "current_user":    current_user,
        "budget":          budget,
        "req":             req,
        "customer":        customer,
        "page_title":      f"Satış Fiyatı — {budget.venue_name}",
        "rows_by_section": rows_by_section,
        "status_label":    BUDGET_STATUS_LABELS.get(budget.budget_status, budget.budget_status),
        "status_color":    BUDGET_STATUS_COLORS.get(budget.budget_status, "secondary"),
        "grouped_services": json.dumps(grouped_services, ensure_ascii=False),
    })


@router.post("/{budget_id}/price", name="budgets_price_save")
async def budgets_price_save(
    budget_id:       str,
    rows_json:       str = Form("[]"),
    service_fee_pct: str = Form("0"),
    manager_notes:   str = Form(""),
    venue_name:      str = Form(""),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("admin", "project_manager"):
        raise HTTPException(403)
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if not budget:
        return RedirectResponse(url="/budgets", status_code=status.HTTP_302_FOUND)

    budget.rows_json       = rows_json
    budget.service_fee_pct = float(service_fee_pct or 0)
    budget.manager_notes   = manager_notes.strip()
    if venue_name.strip():
        budget.venue_name  = venue_name.strip()
    budget.budget_status   = "draft_manager"
    budget.updated_at      = _now()
    db.commit()
    return RedirectResponse(url=f"/budgets/{budget_id}/price", status_code=status.HTTP_302_FOUND)


@router.post("/{budget_id}/approve", name="budgets_approve")
async def budgets_approve(
    budget_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("admin", "project_manager"):
        raise HTTPException(403)
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if budget:
        budget.budget_status = "approved"
        budget.updated_at    = _now()
        req = db.query(ReqModel).filter(ReqModel.id == budget.request_id).first()
        if req:
            req.status     = "completed"
            req.updated_at = _now()
        db.commit()
    return RedirectResponse(url=f"/budgets/{budget_id}", status_code=status.HTTP_302_FOUND)


@router.post("/{budget_id}/request-revision", name="budgets_request_revision")
async def budgets_request_revision(
    budget_id:      str,
    revision_notes: str = Form(""),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("admin", "project_manager"):
        raise HTTPException(403)
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if budget:
        budget.budget_status  = "revision_requested"
        budget.revision_notes = revision_notes.strip()
        budget.updated_at     = _now()
        db.commit()
    return RedirectResponse(url=f"/budgets/{budget_id}", status_code=status.HTTP_302_FOUND)


@router.post("/{budget_id}/cancel", name="budgets_cancel")
async def budgets_cancel(
    budget_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("admin", "project_manager"):
        raise HTTPException(403)
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if budget:
        budget.budget_status = "cancelled"
        budget.updated_at    = _now()
        req = db.query(ReqModel).filter(ReqModel.id == budget.request_id).first()
        if req:
            req.status     = "cancelled"
            req.updated_at = _now()
        db.commit()
    return RedirectResponse(url=f"/budgets/{budget_id}", status_code=status.HTTP_302_FOUND)


@router.get("/{budget_id}/export", name="budgets_export")
async def budgets_export(
    budget_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("admin", "project_manager"):
        raise HTTPException(403)

    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if not budget:
        return RedirectResponse(url="/budgets", status_code=status.HTTP_302_FOUND)

    req      = db.query(ReqModel).filter(ReqModel.id == budget.request_id).first()
    customer = db.query(Customer).filter(Customer.id == req.customer_id).first() if req and req.customer_id else None

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl kurulu değil. pip install openpyxl")

    # ── Template varsa yükle, yoksa boş oluştur ──
    template_path = customer.excel_template_path if customer else ""
    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        config = customer.excel_config if customer else {}
        data_start_row = int(config.get("data_start_row", 1))
        col_map = config.get("columns", {})
    else:
        # Standart format
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Teklif"
        # Başlık satırı
        headers = [
            "Hizmet", "Birim", "Miktar", "Gün/Gece",
            "Birim Fiyat (KDV hariç)", "KDV %",
            "Birim Fiyat (KDV dahil)", "Toplam (KDV dahil)"
        ]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E293B")
            cell.alignment = Alignment(horizontal="center")
        data_start_row = 2
        col_map = {
            "service_name":   "A",
            "unit":           "B",
            "qty":            "C",
            "nights":         "D",
            "sale_price":     "E",
            "vat_rate":       "F",
            "sale_price_inc": "G",
            "total_inc":      "H",
        }

    def col_letter_to_num(letter):
        if not letter:
            return None
        letter = str(letter).upper().strip()
        if letter.isdigit():
            return int(letter)
        from openpyxl.utils import column_index_from_string
        try:
            return column_index_from_string(letter)
        except Exception:
            return None

    SECTION_LABELS = {
        "accommodation": "Konaklama",
        "meeting":       "Toplantı / Salon",
        "fb":            "F&B (Yiyecek & İçecek)",
        "teknik":        "Teknik",
        "dekor":         "Dekor",
        "transfer":      "Transfer",
        "tasarim":       "Tasarım & Baskı",
        "other":         "Diğer",
    }
    SECTIONS_ORDER = ["accommodation", "meeting", "fb", "teknik", "dekor", "transfer", "tasarim", "other"]

    current_row = data_start_row
    rows_by_sec = {}
    for row in budget.rows:
        if row.get("is_service_fee"):
            continue
        sec = row.get("section", "other")
        rows_by_sec.setdefault(sec, []).append(row)

    for sec in SECTIONS_ORDER:
        sec_rows = rows_by_sec.get(sec, [])
        if not sec_rows:
            continue
        # Bölüm başlığı
        if col_map:
            first_col = col_letter_to_num(list(col_map.values())[0])
            last_col  = col_letter_to_num(list(col_map.values())[-1])
            if first_col and last_col:
                ws.cell(row=current_row, column=first_col, value=SECTION_LABELS.get(sec, sec))
                header_cell = ws.cell(row=current_row, column=first_col)
                header_cell.font = Font(bold=True, color="FFFFFF")
                header_cell.fill = PatternFill("solid", fgColor="334155")
                current_row += 1

        for row in sec_rows:
            qty    = float(row.get("qty", 1) or 1)
            nights = float(row.get("nights", 1) or 1)
            sale   = float(row.get("sale_price", 0) or 0)
            vat    = float(row.get("vat_rate", 0) or 0)
            sale_inc  = sale * (1 + vat / 100)
            total_inc = sale_inc * qty * nights

            def wc(field, value, r=current_row):
                col = col_letter_to_num(col_map.get(field))
                if col:
                    ws.cell(row=r, column=col, value=value)

            wc("service_name",   row.get("service_name", ""))
            wc("unit",           row.get("unit", "Adet"))
            wc("qty",            qty)
            wc("nights",         nights)
            wc("sale_price",     sale)
            wc("vat_rate",       f"%{int(vat)}")
            wc("sale_price_inc", round(sale_inc, 2))
            wc("total_inc",      round(total_inc, 2))
            current_row += 1

    # Hizmet bedeli
    sf_row = next((r for r in budget.rows if r.get("is_service_fee")), None)
    if sf_row:
        sf_sale = float(sf_row.get("sale_price", 0) or 0)
        sf_vat  = float(sf_row.get("vat_rate", 20) or 20)
        sf_inc  = sf_sale * (1 + sf_vat / 100)

        def wc_sf(field, value, r=current_row):
            col = col_letter_to_num(col_map.get(field))
            if col:
                ws.cell(row=r, column=col, value=value)

        wc_sf("service_name",   "Hizmet Bedeli")
        wc_sf("unit",           "Hizmet")
        wc_sf("qty",            1)
        wc_sf("nights",         1)
        wc_sf("sale_price",     sf_sale)
        wc_sf("vat_rate",       f"%{int(sf_vat)}")
        wc_sf("sale_price_inc", round(sf_inc, 2))
        wc_sf("total_inc",      round(sf_inc, 2))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = (budget.venue_name or "teklif").replace(" ", "_").replace("/", "-")[:30]
    filename  = f"{safe_name}_teklif.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{budget_id}/delete", name="budgets_delete")
async def budgets_delete(
    budget_id: str,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("admin", "e_dem"):
        raise HTTPException(403)
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if budget and (budget.budget_status == "draft_edem" or current_user.role == "admin"):
        db.delete(budget)
        db.commit()
    return RedirectResponse(url="/budgets", status_code=status.HTTP_302_FOUND)


@router.post("/{budget_id}/copy", name="budgets_copy")
async def budgets_copy(
    budget_id: str,
    rows_json: str = Form(""),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if current_user.role not in ("admin", "project_manager"):
        raise HTTPException(403)
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if not budget:
        raise HTTPException(404)

    # Satış fiyatlarını ve durum bayraklarını kopyala; yeni bütçe pending_manager'dan başlar
    src_rows = rows_json.strip() or budget.rows_json
    new_budget = Budget(
        id=_uuid(),
        request_id=budget.request_id,
        venue_name=budget.venue_name + " (Kopya)",
        rows_json=src_rows,
        budget_status="pending_manager",
        service_fee_pct=budget.service_fee_pct,
        created_by=budget.created_by,
        created_at=_now(),
        updated_at=_now(),
    )
    db.add(new_budget)
    db.commit()
    db.refresh(new_budget)
    return RedirectResponse(url=f"/budgets/{new_budget.id}/price", status_code=status.HTTP_302_FOUND)
