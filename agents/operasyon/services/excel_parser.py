"""
Claude API ile Excel Katılımcı Listesi Parser

Müşterinin gönderdiği (her seferinde farklı formatlı) Excel dosyasını okur,
Claude ile analiz eder ve standart katılımcı formatına dönüştürür.
"""

import json
import os
import time
import openpyxl
from io import BytesIO
from pathlib import Path
from dotenv import load_dotenv
import anthropic

# .env dosyasını yükle (varsa)
load_dotenv(Path(__file__).parent.parent / ".env")

client = anthropic.Anthropic()  # ANTHROPIC_API_KEY env'den okunur

SYSTEM_PROMPT = """Sen bir etkinlik yönetim şirketinin veri analisti asistanısın.
Sana bir Excel dosyasının içeriği verilecek. Bu dosya bir müşterinin gönderdiği
katılımcı listesidir. Formatı standart değil — her müşteri kendi formatını kullanıyor.

Görevin: İçeriği analiz edip her katılımcı için aşağıdaki standart JSON yapısını üretmek.

ÇIKTI FORMATI (JSON array):
[
  {
    "first_name": "string (zorunlu)",
    "last_name": "string (zorunlu)",
    "company": "string veya null",
    "title": "string veya null",
    "email": "string veya null",
    "phone": "string veya null",
    "badge_name": "string veya null (yaka kartı adı, yoksa null)",
    "dietary": "string veya null (beslenme kısıtı)",
    "special_needs": "string veya null",
    "notes": "string veya null",
    "flight_in": {
      "flight_number": "string veya null (örn: TK 2341)",
      "airline": "string veya null",
      "departure_airport": "string veya null (IATA kodu tercihen)",
      "arrival_airport": "string veya null",
      "flight_date": "YYYY-MM-DD veya null",
      "departure_time": "HH:MM veya null",
      "arrival_time": "HH:MM veya null",
      "seat": "string veya null",
      "pnr": "string veya null"
    },
    "flight_out": { ...aynı yapı... },
    "accommodation": {
      "hotel": "string veya null",
      "room_number": "string veya null",
      "room_type": "string veya null (SGL/DBL/SUT vb.)",
      "check_in": "YYYY-MM-DD veya null",
      "check_out": "YYYY-MM-DD veya null",
      "notes": "string veya null"
    }
  }
]

KURALLAR:
- flight_in, flight_out, accommodation alanları: eğer hiç bilgi yoksa null yaz (boş obje değil)
- Tarih formatı mutlaka YYYY-MM-DD olmalı
- Saat formatı mutlaka HH:MM olmalı
- Bilinmeyen/boş alanlar için null kullan
- İsim ayrıştırma: "Ad Soyad" formatındaysa first_name/last_name'e böl
- Sadece JSON array döndür, başka açıklama ekleme
"""


def _read_excel_as_text(content: bytes, filename: str) -> str:
    """Excel dosyasını okunabilir metin formatına çevirir."""
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheets_text = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Tamamen boş satırları atla
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                cells = [str(cell) if cell is not None else "" for cell in row]
                rows.append(" | ".join(cells))

            if rows:
                sheets_text.append(f"=== Sayfa: {sheet_name} ===\n" + "\n".join(rows))

        return "\n\n".join(sheets_text)
    except Exception as e:
        raise ValueError(f"Excel dosyası okunamadı: {e}")


async def parse_participant_excel(content: bytes, filename: str) -> list[dict]:
    """
    Excel içeriğini Claude ile analiz eder, standart katılımcı listesi döner.
    Overload (529) durumunda 3 kez yeniden dener.
    """
    excel_text = _read_excel_as_text(content, filename)

    if not excel_text.strip():
        raise ValueError("Excel dosyası boş veya okunamadı.")

    # Çok büyükse kırp (Claude'un bağlam sınırı için)
    MAX_CHARS = 60_000
    if len(excel_text) > MAX_CHARS:
        excel_text = excel_text[:MAX_CHARS] + "\n\n[... dosya kırpıldı, ilk kısım işleniyor ...]"

    last_error = None
    for attempt in range(3):
        try:
            message = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=8096,
                system=SYSTEM_PROMPT,
                messages=[
                    {
                        "role": "user",
                        "content": f"Dosya adı: {filename}\n\nİçerik:\n{excel_text}"
                    }
                ]
            )
            break  # Başarılı
        except anthropic.APIStatusError as e:
            last_error = e
            if e.status_code in (529, 503, 529):
                wait = 5 * (attempt + 1)
                time.sleep(wait)
                continue
            raise ValueError(f"Claude API hatası ({e.status_code}): {e.message}")
        except anthropic.AuthenticationError:
            raise ValueError("ANTHROPIC_API_KEY geçersiz veya eksik. Lütfen .env dosyasını kontrol edin.")
    else:
        raise ValueError(
            f"Claude API şu an yoğun (overloaded). Birkaç dakika sonra tekrar deneyin. "
            f"Hata: {last_error}"
        )

    response_text = message.content[0].text.strip()
    return _extract_json_array(response_text)


def _extract_json_array(text: str) -> list[dict]:
    """
    Claude'un çıktısından JSON array'i çıkarır.
    Farklı formatları dener: düz JSON, ```json blok, metin içine gömülü [...].
    """
    # 1. Düz JSON array
    if text.startswith("["):
        try:
            result = json.loads(text)
            if isinstance(result, list):
                return result
        except json.JSONDecodeError:
            pass

    # 2. ```json ... ``` veya ``` ... ``` bloğu
    import re
    code_block = re.search(r"```(?:json)?\s*(\[[\s\S]*?\])\s*```", text)
    if code_block:
        try:
            result = json.loads(code_block.group(1))
            if isinstance(result, list):
                return result
        except json.JSONDecodeError:
            pass

    # 3. Metin içinde ilk [ ... ] bloğunu bul
    start = text.find("[")
    if start != -1:
        # Eşleşen ] bul (iç içe sayarak)
        depth = 0
        for i, ch in enumerate(text[start:], start):
            if ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    candidate = text[start:i+1]
                    try:
                        result = json.loads(candidate)
                        if isinstance(result, list):
                            return result
                    except json.JSONDecodeError:
                        break

    raise ValueError(
        "Claude geçerli bir JSON array döndürmedi. "
        "Lütfen tekrar deneyin."
    )
